'''
    Automação
'''

import os
import datetime
from openpyxl import load_workbook, Workbook

# Funcionando

final_wb = Workbook()
final_ws = final_wb.active
final_ws.title = "sonepar_support_report"

yesterday=(datetime.date.today() - datetime.timedelta(days=1))
if yesterday.strftime("%A") == "Sunday":
    yesterday=(datetime.date.today() - datetime.timedelta(days=3))

FOLDER_PATH = "C:/Users/P0589/Downloads/Suporte_Sonepar"
folder = os.listdir(FOLDER_PATH)
for file in folder:
    file_path = os.path.join(FOLDER_PATH, file)
    if "task_sla" in file:
        bulk_sla = load_workbook(filename= file_path, data_only=True).active
    if "incident" in file:
        incident_data = load_workbook(filename= file_path, data_only=True).active
    if "sc_req_item" in file:
        requisiton_data = load_workbook(filename= file_path, data_only=True).active

# Título das colunas
titles = ["Dia",
          "ID Chamado",
          "Sistema",
          "Cliente",
          "Status",
          "Descrição",
          "Atendente",
          "Data Fechamento",
          "Aguardando",
          "Motivo",
          "SLA de Resolução %",
          "SLA de Resolução Tempo"
        ]

# Preenche os títulos nas colunas da planilha final
for x, title in enumerate(titles):
    final_ws.cell(row=1, column= x+1, value=title)

for id_inc in incident_data["B"][1:]: # Percorre todos os IDs de chamados na Coluna B da planilha
    new_row = final_ws.max_row+1 # Próxima linha vazia a ser inserido os dados
    COLUMN=1
    # Preenche os dados das Colunas A, B, C D da planilha de incidentes na planilha final
    for row in incident_data[f"A{id_inc.row}:D{id_inc.row}"]:
        for cell in row:
            final_ws.cell(row=new_row, column=COLUMN, value=cell.value)
            COLUMN+=1
    if incident_data[f"G{id_inc.row}"].value is None: # Faz a verificação do Status
        final_ws.cell(row=new_row, column=COLUMN, value="Aberto")
        COLUMN+=1
    else:
        final_ws.cell(row=new_row, column=COLUMN, value="Fechado")
        COLUMN+=1
    # Preenche os dados das Colunas E, F, G, H, I da planilha de incidentes na planilha
    for row in incident_data[f"E{id_inc.row}:I{id_inc.row}"]:
        for cell in row:
            final_ws.cell(row=new_row, column=COLUMN, value=cell.value)
            COLUMN+=1
    for id_chamado in bulk_sla["A"][1:]: # Procura o mesmo Id de chamado dentro da planilha de SLA
        if id_chamado.value == id_inc.value:
            # Insere o valor da porcentagem na planilha final
            final_ws.cell(row=new_row, column=11,
                          value=bulk_sla[f"I{id_chamado.row}"].value)
            # Faz o cálculo do tempo e insere na planilha final
            final_ws.cell(row=new_row, column=12,
                          value=bulk_sla[f"H{id_chamado.row}"].value/86400)

# Faz a mesma coisa só que para os dados de Itens Requisitados
for id_ritm in requisiton_data["B"][1:]:
    new_row = final_ws.max_row+1 # Próxima linha vazia a ser inserido os dados
    COLUMN=1
    for row in requisiton_data[f"A{id_ritm.row}:D{id_ritm.row}"]:
        for cell in row:
            final_ws.cell(row=new_row, column=COLUMN, value=cell.value)
            COLUMN+=1
    if requisiton_data[f"G{id_ritm.row}"].value is None:
        final_ws.cell(row=new_row, column=COLUMN, value="Aberto")
        COLUMN+=1
    else:
        final_ws.cell(row=new_row, column=COLUMN, value="Fechado")
        COLUMN+=1
    for row in requisiton_data[f"E{id_ritm.row}:H{id_ritm.row}"]:
        for cell in row:
            final_ws.cell(row=new_row, column=COLUMN, value=cell.value)
            COLUMN+=1
    for id_chamado in bulk_sla["A"][1:]:
        if id_chamado.value == id_ritm.value:
            final_ws.cell(row=new_row, column=11, value=bulk_sla[f"I{id_chamado.row}"].value)
            final_ws.cell(row=new_row, column=12, value=bulk_sla[f"H{id_chamado.row}"].value/86400)

#os.remove(f"{FOLDER_PATH}/task_sla.xlsx")
#os.remove(f"{FOLDER_PATH}/incident.xlsx")
#os.remove(f"{FOLDER_PATH}/sc_req_item.xlsx")

final_wb.save(f"{FOLDER_PATH}/suporte_sonepar_att_{yesterday}.xlsx")
print("Processo Finalizado!")
