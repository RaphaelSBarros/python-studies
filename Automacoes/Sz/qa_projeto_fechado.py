'''
    Automação
'''
import os
import datetime
from openpyxl import load_workbook, Workbook
import pandas as pd

# Testar na proxima atualização
FOLDER_PATH = "C:/Users/P0589/Downloads/QA_Projeto_Fechado"
folder = os.listdir(FOLDER_PATH)
inicio1 = datetime.datetime.now()
final_wb = Workbook()
final_ws = final_wb.active
final_ws.title = "qa_projeto_fechado"

yesterday=(datetime.date.today() - datetime.timedelta(days=1))
if yesterday.strftime("%A") == "Sunday":
    yesterday=(datetime.date.today() - datetime.timedelta(days=3))

REPORT = None
UPDATE = None
for file in folder:
    file_path = os.path.join(FOLDER_PATH, file)
    if "Relatório" in file:
        b = pd.read_excel(io=file_path, sheet_name="QA_PROJETO_FECHADO")
        CONVERTED="data/CONVERTED.xlsx"
        b.to_excel(CONVERTED, index=False)
        REPORT = load_workbook(filename=CONVERTED, data_only=True).active
        os.remove(CONVERTED)
    if "jira" in file:
        UPDATE = load_workbook(filename=file_path, data_only=True).active

titles = ["Dia",
          "ID Atividade",
          "Sistema",
          "Tipo",
          "Status",
          "Descrição",
          "ATENDENTE",
          "Data de Fechamento"
        ]
for x, title in enumerate(titles):
    final_ws.cell(row=1, column=x+1, value=title)

## Atualizar chamados antigos
existentes = [] # Armazenando quais os chamados existentes no relatório
UPDATE_COUNTER=0 # Contagem dos chamados com atualização
for chamado in REPORT["B"][2:]:
    if chamado.value is None:
        break
    existentes.append(chamado.value)
    new_row = final_ws.max_row+1
    # Preenchimento dos dados existentes na tabela final
    final_ws.cell(row=new_row, column=1, value=REPORT[f"A{chamado.row}"].value)
    final_ws.cell(row=new_row, column=2, value=chamado.value)
    final_ws.cell(row=new_row, column=3, value=REPORT[f"C{chamado.row}"].value)
    final_ws.cell(row=new_row, column=4, value=REPORT[f"D{chamado.row}"].value)
    final_ws.cell(row=new_row, column=5, value=REPORT[f"E{chamado.row}"].value)
    final_ws.cell(row=new_row, column=6, value=REPORT[f"F{chamado.row}"].value)
    final_ws.cell(row=new_row, column=7, value=REPORT[f"G{chamado.row}"].value)
    final_ws.cell(row=new_row, column=8, value=REPORT[f"H{chamado.row}"].value)
    # Trocando os dados já existentes com os atualizados pelo relatório jira
    for atualizar in UPDATE["A"][1:]:
        if chamado.value == atualizar.value:
            # Verificação do status para se encaixar no padrão da planilha KPIs
            if UPDATE[f"F{atualizar.row}"].value in ("Concluído", "DEPLOY"):
                final_ws.cell(row=new_row, column=5, value="Fechado")
                final_ws.cell(row=new_row, column=8, value=UPDATE[f"O{atualizar.row}"].value)
            else:
                final_ws.cell(row=new_row, column=5, value="Aberto")
            UPDATE_COUNTER+=1

# Adicionar Novos Chamados
ADD_COUNTER = 0 # Cotagem dos novos chamados a serem inseridos
for novos in UPDATE["A"][1:]:
    if novos.value not in existentes: # Inserindo os novos dados
        new_row = final_ws.max_row+1
        final_ws.cell(row=new_row, column=1, value=yesterday)
        final_ws.cell(row=new_row, column=2, value=novos.value)
        final_ws.cell(row=new_row, column=3, value=UPDATE[f"D{novos.row}"].value)
        final_ws.cell(row=new_row, column=4, value=UPDATE[f"E{novos.row}"].value)
        if UPDATE[f"F{novos.row}"].value in ("Concluído", "DEPLOY"):
            final_ws.cell(row=new_row, column=5, value="Fechado")
        else:
            final_ws.cell(row=new_row, column=5, value="Aberto")
        final_ws.cell(row=new_row, column=6, value=UPDATE[f"B{novos.row}"].value)
        final_ws.cell(row=new_row, column=7, value=UPDATE[f"M{novos.row}"].value)
        final_ws.cell(row=new_row, column=8, value=UPDATE[f"O{novos.row}"].value)
        ADD_COUNTER+=1

# Salva a nova tabela na mesma pasta que estavam os relatórios
final_wb.save(f"data/qa_projeto_fechado_att_{yesterday}.xlsx")
# Apenas para conferência do total de informações
print(f"Projeto Finalizado com {UPDATE_COUNTER} atualizações e {ADD_COUNTER} adições nos dados")
