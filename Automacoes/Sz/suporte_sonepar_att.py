'''
    Automação
'''

import datetime
import os

from openpyxl import load_workbook, Workbook
import pandas as pd

FOLDER_PATH = "C:/Users/P0589/Downloads/Suporte_Sonepar"
folder = os.listdir(FOLDER_PATH)

final_wb = Workbook()
final_ws = final_wb.active

yesterday=(datetime.date.today() - datetime.timedelta(days=1))
if yesterday.strftime("%A") == "Sunday":
    yesterday=(datetime.date.today() - datetime.timedelta(days=3))

REPORT = None
UPDATE = None
for file in folder:
    file_path = os.path.join(FOLDER_PATH, file)
    if "Relatório" in file:
        b = pd.read_excel(io=file_path, sheet_name="SUPORTE_SONEPAR")
        CONVERTED="data/CONVERTED.xlsx"
        b.to_excel(CONVERTED, index=False)
        REPORT = load_workbook(filename=CONVERTED, data_only=True).active
        os.remove(CONVERTED)
    if "att" in file:
        UPDATE = load_workbook(filename=file_path, data_only=True).active

titles = [
          "Dia",
          "ID Chamado",
          "Sistema",
          "Cliente",
          "Categoria",
          "Status",
          "Descrição",
          "Atendente",
          "Data Fechamento",
          "Aguardando",
          "Motivo",
          "SLA de Resolução %",
          "SLA de Resolução Tempo"
        ] # Título das colunas

# Preenche os títulos nas colunas da planilha final
for x, title in enumerate(titles):
    final_ws.cell(row=1, column=x+1, value=title)

existentes = [] # Armazenando quais os chamados existentes no relatório
UPDATE_COUNTER=0 # Contagem dos chamados com atualização
for chamado in REPORT["B"][2:]:
    if chamado.value is None:
        break
    existentes.append(chamado.value)
    new_row = final_ws.max_row+1
    for row in final_ws[f"A{new_row}:M{new_row}"]:
        for cell in row:
            final_ws.cell(row=new_row,
                          column=cell.column,
                          value=REPORT.cell(row=chamado.row, column=cell.column).value)
    for atualizar in UPDATE["B"][1:]:
        if (atualizar.value == chamado.value
            and
            UPDATE[f"E{atualizar.row}"].value != REPORT[f"F{chamado.row}"].value):
            UPDATE_COUNTER+=1
            final_ws.cell(row=new_row, column=6, value=UPDATE[f"E{atualizar.row}"].value)
            final_ws.cell(row=new_row, column=9, value=UPDATE[f"G{atualizar.row}"].value)
            final_ws.cell(row=new_row, column=10, value=UPDATE[f"H{atualizar.row}"].value)
            final_ws.cell(row=new_row, column=11, value=UPDATE[f"I{atualizar.row}"].value)
            final_ws.cell(row=new_row, column=12, value=UPDATE[f"J{atualizar.row}"].value)
            final_ws.cell(row=new_row, column=13, value=UPDATE[f"K{atualizar.row}"].value)
            print(f"Chamado Atualizado: {chamado.value}")

COUNTER = 0
for novos in UPDATE["B"][1:]:
    if novos.value not in existentes:
        COUNTER+=1
        new_row = final_ws.max_row+1
        for row in final_ws[f"A{new_row}:D{new_row}"]:
            for cell in row:
                final_ws.cell(row=new_row,
                              column=cell.column,
                              value=UPDATE.cell(row=novos.row, column=cell.column).value)
        for row in final_ws[f"F{new_row}:M{new_row}"]:
            for cell in row:
                final_ws.cell(row=new_row,
                              column=cell.column,
                              value=UPDATE.cell(row=novos.row, column=cell.column-1).value)

# Salva a nova tabela na mesma pasta que estavam os relatórios
final_wb.save(f"{FOLDER_PATH}/atualizacao_sonepar.xlsx")

# Apenas para conferência do total de informações
print(f"Projeto Finalizado com:\n{UPDATE_COUNTER} atualizações e {COUNTER} adições nos dados")
