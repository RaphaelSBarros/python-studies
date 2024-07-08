'''
    Extração dos RPWs e alocação em linhas de planilhas
'''

from tkinter import filedialog
from openpyxl import load_workbook, Workbook
#import pandas as pd


file_path = filedialog.askopenfilename()
rpw_data = load_workbook(filename=file_path, data_only=True)['RPW']

wb = Workbook()
ws = wb.active
ws.title = "RPW"

titles = ["Data",
          "RPW Afetado",
          "Empresa",
          "Resolução",
          ]

for x, title in enumerate(titles):
    ws.cell(row=1, column=x+1, value=title)

for row in rpw_data["B"][1:]:
    if row.value is None:
        break
    if row.value != 0 and "," in row.value:
        rpws = row.value.replace(",", "").split(" ")
        for rpw in rpws:
            if len(rpw) > 1:
                new_row = ws.max_row+1
                ws.cell(row=new_row, column=1, value=rpw_data[f"A{row.row}"].value)
                ws.cell(row=new_row, column=2, value=rpw)
                if "NOR" in rpw:
                    ws.cell(row=new_row, column=3, value="Nortel")
                else:
                    ws.cell(row=new_row, column=3, value="Dimensional")
                ws.cell(row=new_row, column=4, value=rpw_data[f"C{row.row}"].value)

wb.save("data/rpw.xlsx")
