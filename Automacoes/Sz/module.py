import os
from openpyxl import load_workbook, Workbook
import pandas as pd
import datetime
from tkinter import filedialog

class Automatization:
    def __init__(self, wb_title: str = "Sheet"):
        self.wb = Workbook()
        self.wb.active.title = wb_title

    def create_worksheet(self, *worksheets: str):
        for ws in worksheets:
            self.wb.create_sheet(ws)

    def fill_worksheet(self, worksheet, row: int, *values):
        for x, value in enumerate(values):
            self.wb[worksheet].cell(row=row,column=x+1,value= value)

    @classmethod
    def csv_to_excel(self, ):
        file = filedialog.askopenfilename()
        NEW_FILE = "convfile.xlsx"
        pd.read_csv(file, delimiter=";").to_excel(NEW_FILE, index=False)

    def save_workbook(self, filename: str):
        self.wb.save(f"{filename}.xlsx")

    def get_last_businessday(self):
        yesterday=(datetime.date.today() - datetime.timedelta(days=1))
        if yesterday.strftime("%A") == "Sunday":
            yesterday=(datetime.date.today() - datetime.timedelta(days=3))
        return yesterday
    
    def entitle_worksheet(self, titles, sheetname: str):
        for x, title in enumerate(titles):
            self.wb[sheetname].cell(row=1, column=x+1, value=title)

    def str_to_datetime(self, string_date):
        iD,iM,iY, = [int(x) for x in string_date.split("/")]
        return datetime.datetime(iY,iM,iD)
