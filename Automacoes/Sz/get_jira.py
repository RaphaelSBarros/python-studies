"""Automação
"""

import os
from openpyxl import load_workbook, Workbook
import pandas as pd

FOLDER_PATH = "C:/Users/P0589/Downloads/QA_Projeto_Fechado"
folder = os.listdir(FOLDER_PATH)
final_wb = Workbook()
final_ws = final_wb.active
final_ws.title = "ID_Jira"

for file in folder:
    file_path = os.path.join(FOLDER_PATH, file)
    if "Relatório" in file:
        b = pd.read_excel(io=file_path, sheet_name="SUPORTE_CLIENTE_EXTERNO")
        CONVERTED="data/CONVERTED.xlsx"
        b.to_excel(CONVERTED, index=False)
        REPORT_WB = load_workbook(filename=CONVERTED, data_only=True)
        REPORT = REPORT_WB.active
        os.remove(CONVERTED)
    if "jira" in file:
        UPDATE = load_workbook(filename=file_path, data_only=True).active
ADDED=0
for chamado in REPORT["B"][2:]:
    JIRA=[]
    for nr_chamado in UPDATE["Q"][1:]:
        if str(nr_chamado.value) == str(chamado.value):
            JIRA.append(UPDATE[f"A{nr_chamado.row}"].value)
            ADDED+=1
    REPORT[f"C{chamado.row}"].value = ", ".join(JIRA)
REPORT_WB.save("teste.xlsx")
print(f"Atualização finalizada com {ADDED} jiras identificados")
