'''
    Automação
'''

from tkinter import filedialog
import datetime
import os
from openpyxl import load_workbook, Workbook
import pandas as pd

# Precisa automatizar isso?

file_path = filedialog.askopenfilename()
df = pd.read_csv(file_path, delimiter=";")
NEW_FILE = "data/convfile.xlsx"
df.to_excel(NEW_FILE, index=False)
relatorio = load_workbook(NEW_FILE, data_only=True).active
os.remove(NEW_FILE)

final_wb = Workbook()
int_final_ws = final_wb.active
int_final_ws.title = "suporte_cliente_interno"
final_wb.create_sheet("suporte_cliente_externo")
ext_final_ws = final_wb["suporte_cliente_externo"]

int_titles =[
        "Dia", 
        "Chamado", 
        "Sistema", 
        "Area", 
        "Categoria", 
        "Status", 
        "Atendente", 
        "Descrição", 
        "Data de Fechamento", 
        "SLA Primeiro Atendimento"
    ]
ext_titles =[
        "Dia", 
        "Chamado", 
        "Sistema", 
        "Cliente", 
        "Categoria", 
        "Status", 
        "Descrição", 
        "Atendente", 
        "Data de Fechamento", 
        "SLA Primeiro Atendimento"
    ]

yesterday=(datetime.date.today() - datetime.timedelta(days=1))
if yesterday.strftime("%A") == "Sunday":
    yesterday=(datetime.date.today() - datetime.timedelta(days=3))

for x, title in enumerate(int_titles):
    int_final_ws.cell(row=1, column=x+1, value=title)

for x, title in enumerate(ext_titles):
    ext_final_ws.cell(row=1, column=x+1, value=title)

for cliente in relatorio["C"][1:]:
    int_new_row = int_final_ws.max_row+1
    ext_new_row = ext_final_ws.max_row+1

    itime = relatorio[f"O{cliente.row}"].value.split(":")
    ftime = relatorio[f"N{cliente.row}"].value.split(":")

    idate = relatorio[f"M{cliente.row}"].value.split("/")
    fdate = relatorio[f"L{cliente.row}"].value.split("/")

    if cliente.value == "SZ Soluções":
        int_final_ws.cell(row=int_new_row, column=1, value=yesterday)
        int_final_ws.cell(row=int_new_row, column=2, value= relatorio[f"A{cliente.row}"].value)
        int_final_ws.cell(row=int_new_row, column=3, value= cliente.value)
        int_final_ws.cell(row=int_new_row, column=4, value= relatorio[f"E{cliente.row}"].value)
        int_final_ws.cell(row=int_new_row, column=5, value= relatorio[f"H{cliente.row}"].value)
        if relatorio[f"F{cliente.row}"].value in ("Encerrado"):
            int_final_ws.cell(row=int_new_row, column=6, value="Fechado")
        else:
            int_final_ws.cell(row=int_new_row, column=6, value="Aberto")
        int_final_ws.cell(row=int_new_row, column=7, value= relatorio[f"D{cliente.row}"].value)
        int_final_ws.cell(row=int_new_row, column=8, value= relatorio[f"B{cliente.row}"].value)
        int_final_ws.cell(row=int_new_row, column=9, value= relatorio[f"G{cliente.row}"].value)
        int_final_ws.cell(row=int_new_row, column=10,
            value= datetime.datetime(
                year=int(fdate[2]),
                month=int(fdate[1]),
                day=int(fdate[0]),
                hour=int(ftime[0]),
                minute=int(ftime[1]),
                second=int(ftime[2])
            ) -
            datetime.datetime(
                year=int(idate[2]),
                month=int(idate[1]),
                day=int(idate[0]),
                hour=int(itime[0]),
                minute=int(itime[1]),
                second=int(itime[2])
            )
        )
    else:
        ext_final_ws.cell(row=ext_new_row, column=1, value=yesterday)
        ext_final_ws.cell(row=ext_new_row, column=2, value= relatorio[f"A{cliente.row}"].value)
        if "Sonepar" in cliente.value:
            split = [x for x in cliente.value.split(" ") if len(x) > 1]
            ext_final_ws.cell(row=ext_new_row, column=3, value= split[0])
            ext_final_ws.cell(row=ext_new_row, column=4, value= split[1])
        else:
            ext_final_ws.cell(row=ext_new_row, column=3, value="BeeStock")
            ext_final_ws.cell(row=ext_new_row, column=4, value= relatorio[f"C{cliente.row}"].value)
        ext_final_ws.cell(row=ext_new_row, column=5, value= relatorio[f"E{cliente.row}"].value)
        if relatorio[f"F{cliente.row}"].value in ("Encerrado"):
            ext_final_ws.cell(row=ext_new_row, column=6, value="Fechado")
        else:
            ext_final_ws.cell(row=ext_new_row, column=6, value="Aberto")
        ext_final_ws.cell(row=ext_new_row, column=7, value= relatorio[f"B{cliente.row}"].value)
        ext_final_ws.cell(row=ext_new_row, column=8, value= relatorio[f"D{cliente.row}"].value)
        ext_final_ws.cell(row=ext_new_row, column=9, value= relatorio[f"G{cliente.row}"].value)
        ext_final_ws.cell(row=ext_new_row, column=10,
            value =
            datetime.datetime(
                year=int(fdate[2]),
                month=int(fdate[1]),
                day=int(fdate[0]),
                hour=int(ftime[0]),
                minute=int(ftime[1]),
                second=int(ftime[2])
            ) -
            datetime.datetime(
                year=int(idate[2]),
                month=int(idate[1]),
                day=int(idate[0]),
                hour=int(itime[0]),
                minute=int(itime[1]),
                second=int(itime[2])
            )
        )

final_wb.save("data/suporte_cliente.xlsx")
