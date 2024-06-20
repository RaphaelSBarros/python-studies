from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import pandas as pd
import datetime
import os

# Funcionando

folder_path = "C:/Users/P0589/Downloads/QA_Outsourcing"
folder = os.listdir(folder_path)

final_wb = Workbook()
final_ws = final_wb.active
final_ws.title = "QA_Outsourcing"

titles =["Período", "ID Atividade", "Projeto", "Tipo", "Status", "Descrição", "Atendente", "Data de Fechamento"]

yesterday=(datetime.date.today() - datetime.timedelta(days=1))
if yesterday.strftime("%A") == "Sunday":
    yesterday=(datetime.date.today() - datetime.timedelta(days=3))

for x in range(len(titles)):
    final_ws.cell(row=1, column=x+1, value=titles[x])

for file in folder:
    if ".csv" in file:
        file_path = os.path.join(folder_path, file)
        df = pd.read_csv(file_path)
        xlsx_path = f"data/{file[:-3]}xlsx"
        df.to_excel(xlsx_path, index=False)
        ws = load_workbook(filename=xlsx_path, data_only=True).active
        for filter in ws["D"][1:]:
            if filter.value in ("Pronto para Homologar", "Pronto para Validar", "Validación", "Homologación"):
                new_row = final_ws.max_row+1
                final_ws.cell(row=new_row, column=1, value=yesterday)
                for row in ws[f"A{filter.row}:G{filter.row}"]:
                    for x, cell in enumerate(row):
                        if x == 1:
                            if "Alianza" in cell.value:
                                final_ws.cell(row=new_row, column=x+2, value="ALIANZA")
                            elif "Checkbuy Brasil" in cell.value:
                                final_ws.cell(row=new_row, column=x+2, value="CHECKBUY BRASIL")
                            elif "Portal de Compras" in cell.value and "DIMENSIONAL" in cell.value:
                                final_ws.cell(row=new_row, column=x+2, value=r"COMPRAS\DIMENSIONAL")
                            elif "Portal de Compras" in cell.value and "NORTEL" in cell.value:
                                final_ws.cell(row=new_row, column=x+2, value=r"COMPRAS\NORTEL")
                            elif "Datasul_Dimensional" in cell.value:
                                final_ws.cell(row=new_row, column=x+2, value=r"DATASUL\DIMENSIONAL")
                            elif "Datasul_Nortel" in cell.value:
                                final_ws.cell(row=new_row, column=x+2, value=r"DATASUL\NORTEL")
                            elif "Sphere - Nortel" in cell.value:
                                final_ws.cell(row=new_row, column=x+2, value=r"SPHERE\NORTEL")
                            elif "Sphere - Perú" in cell.value:
                                final_ws.cell(row=new_row, column=x+2, value=r"SPHERE\PERU")
                            else:
                                final_ws.cell(row=new_row, column=x+2, value=cell.value.upper())
                        else:
                            final_ws.cell(row=new_row, column=x+2, value=cell.value)               
        os.remove(xlsx_path)
final_wb.save(f"data/qa_outsourcing_att_{yesterday}.xlsx")
print("Processo Finalizado!")