import time
from pikepdf import Pdf
from openpyxl import load_workbook
import os

#///////////////////////////// UTILIZAR O MODELO CRIADO PARA PREENCHER EMAILS /////////////////////////////

def fill_mensalists_data(workbook):
    mensalistas = [] # Cria o array das listas com os dados de cada mensalista
    wb = load_workbook(filename=workbook, data_only= True) # Carrega a planilha com as informações das admissões
    ws = wb.active # Escolhe a aba ativa da planilha
    conf=[] # Cria uma lista que armazenará os REs dos mensalistas para que não haja duplicidade nos dados
    for type in ws['AR'][1:]:
        if type.value == "MENSAL.":
            nome = ws[f'B{type.row}'].value
            re = ws[f'A{type.row}'].value
            email = ws[f'Q{type.row}'].value
            if re not in conf: # Utilizando o array para evitar duplicidade
                mensalistas.append({'nome': nome, 're': re, 'email': email}) # Armazena os dados na lista
                conf.append(re)
    return mensalistas

def slice_and_rename_from_data(file, mensalists):
    with Pdf.open(file) as pdf:
        if (len(pdf.pages)) % (len(mensalists)) == 0: # Verificação para que o número de contratos seja compatível com o número de mensalistas
            div = int(len(pdf.pages) / len(mensalists)) # Armazena o número de páginas por contrato
            for i in range(0, len(pdf.pages), div):
                new = Pdf.new() # Cria um novo arquivo pdf
                for j in range(div): # Insere as páginas do contrato nesse novo arquivo pdf
                    new.pages.append(pdf.pages[i+j])
                title = (f'CONTRATO DE TRABALHO - {mensalists[int(i/div)]['nome']} - {mensalists[int(i/div)]['re']}.pdf') # Renomeando o arquivo no padrão solicitado
                savefile = os.path.join(r'C:\Users\DESOUR10\Downloads\pedeefi', title) # Armazenando o arquivo na pasta especificada
                new.save(savefile) # Salva o arquivo e finaliza o processo
        else:
            print("Número de colaboradores incompatível com o número de contratos!")
