from tkinter import filedialog
from openpyxl import load_workbook
import os
import datetime

folder_path = filedialog.askdirectory() # Escolhe o diretório do arquivo
folder = os.listdir(folder_path)

conferencia=''
contas=''
dependentes=''
eSocial=''

for file in folder:
    if 'Conferência' in file:
        file_path = os.path.join(folder_path, file)
        conferencia = load_workbook(filename=file_path, data_only= True)
        conferencia = conferencia.active
        print('Conferência ok')
    elif 'Contas' in file:
        file_path = os.path.join(folder_path, file)
        contas = load_workbook(filename=file_path, data_only= True)
        contas = contas.active
        print('Contas ok')
    elif 'Dependente' in file:
        file_path = os.path.join(folder_path, file)
        dependentes = load_workbook(filename=file_path, data_only= True)
        dependentes = dependentes.active
        print('Dependentes ok')
    elif 'eSocial' in file:
        file_path = os.path.join(folder_path, file)
        eSocial = load_workbook(filename=file_path, data_only= True)
        eSocial = eSocial.active
        print('eSocial ok')
    elif 'WorkForce' in file:
        file_path = os.path.join(folder_path, file)
        workForce = load_workbook(filename=file_path, data_only= True)
        workForce = workForce.active
        print('WorkForce ok')
    elif 'Check' in file:
        file_path = os.path.join(folder_path, file)
        checkList = load_workbook(filename=file_path)
        check = checkList
        checkList = checkList['Conferência']
        print('CheckList ok')
    else:
        print("erro")

for re in checkList['B'][2:]:
    resConferidos = []
    nomesConferidos=[]
    if re.value is not None:
        for celula in conferencia['A'][1:]:
            linha = celula.row
            ## Verifica se o re informado está dentro da planilha
            if celula.value == re.value:
                # Verificação para evitar duplicidades nos dados
                if re.value not in resConferidos:
                    ## Formatando as datas
                    dtInicio = str(conferencia[f'C{linha}'].value)
                    dtInicio = datetime.date(int(dtInicio[:4]),int(dtInicio[5:7]),int(dtInicio[8:10]))
                    vinculo = conferencia[f'AF{linha}'].value
                    codCategoria = conferencia[f'AL{linha}'].value
                    codVinculo = conferencia[f'AM{linha}'].value
                    codExposicao = conferencia[f'AN{linha}'].value
                    ##### Agência e Conta #####
                    agencia='-'
                    conta='-'
                    for cell in contas['A'][1:]:
                        if cell.value == re.value:
                            row = cell.row
                            agencia = str(contas[f'D{row}'].value)
                            agencia = agencia[-4:] ##Agência
                            conta = contas[f'E{row}'].value##Conta
                    dependente = 0
                    for cell in dependentes['A'][1:]:
                        row = cell.row
                        if cell.value == re.value:
                            nome = dependentes[f'L{row}'].value
                            if nome not in nomesConferidos:
                                dependente+=1
                                nomesConferidos.append(nome)
                    saude = conferencia[f'P{linha}'].value
                    saudeDt = str(conferencia[f'S{linha}'].value)
                    saudeDt = datetime.date(int(saudeDt[:4]),int(saudeDt[5:7]),int(saudeDt[8:10]))
                    vida = conferencia[f'V{linha}'].value
                    vidaDt = str(conferencia[f'X{linha}'].value)
                    vidaDt = datetime.date(int(vidaDt[:4]),int(vidaDt[5:7]),int(vidaDt[8:10]))
                    fgts = conferencia[f'AG{linha}'].value
                    pFgts = conferencia[f'AH{linha}'].value
                    infotipo = '-'
                    turno = conferencia[f'K{linha}'].value
                    ## Formatando a carga horária
                    cargaHoraria = str(conferencia[f'J{linha}'].value)
                    cargaHoraria = f'{cargaHoraria[:-2]} hs'
                    salario = conferencia[f'AE{linha}'].value
                    sindicato = conferencia[f'AB{linha}'].value
                    gremio='-'
                    if conferencia[f'A{linha}'].value == conferencia[f'A{linha+1}'].value:
                        gremio = conferencia[f'AB{linha+1}'].value
                    periculosidade = conferencia[f'AK{linha}'].value
                    es='-'
                    for cell in eSocial['J'][1:]:
                        row = cell.row
                        if re.value in cell.value:
                            es = eSocial[f'G{row}'].value
                    wf='-'
                    cargo=''
                    for cell in workForce['BD'][1:]:
                        row = cell.row
                        if int(re.value) == cell.value:
                            wf = workForce[f'CB{row}'].value
                            cargo = workForce[f'BJ{row}'].value
                    varArray = [
                         dtInicio,
                         cargo,
                         vinculo,
                         codCategoria,
                         codVinculo,
                         codExposicao,
                         agencia,
                         conta,
                         dependente,
                         saude,
                         saudeDt,
                         vida,
                         vidaDt,
                         fgts,
                         pFgts,
                         infotipo,
                         turno,
                         cargaHoraria,
                         salario,
                         gremio,
                         sindicato,
                         periculosidade,
                         es,
                         wf
                    ]
                    for col in checkList[f'AC{re.row}':f'AZ{re.row}']:
                        for i, cell in enumerate(col):
                            cell.value = varArray[i]
                    resConferidos.append(re.value)
## Salva o arquivo com a data atual
now = str(datetime.datetime.now())
filename = f'Att-{now[8:10]}-{now[5:7]}.xlsx'
savefile = os.path.join(folder_path, filename)
check.save(f"{savefile}")